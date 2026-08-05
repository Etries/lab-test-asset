"""
make_template.py  -  regenerate the import template + a sample dataset that
match the real column schema (from config.py) exactly.
Produces: import_template.xlsx (blank-ish, with a few example rows)
          sample_data.xlsx      (fuller sample to load and click around)
Run:  python make_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import random, datetime
import config
random.seed(11)

HDR = Font(bold=True, color="FFFFFF"); FILL = PatternFill("solid", fgColor="1F4E79")
EX = Font(italic=True, color="808080")

SITES = ["Ballinakill", "Bagenalstown", "Maidenhead", "Gorteen",
         "Ballygunteen", "Bolthu Hill", "Stock", "Recycled"]
REAL_SITES = SITES[:6]  # actual physical sites (not Stock/Recycled)

def battery_headers():
    return [f["key"] for f in config.BATTERY_FIELDS]

def new_wb():
    wb = openpyxl.Workbook(); del wb["Sheet"]; return wb

def add_sheet(wb, title, headers):
    ws = wb.create_sheet(title)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.font = HDR; c.fill = FILL
    ws.freeze_panes = "A2"
    return ws

def blankish(v, p): return None if random.random() < p else v

# ---------- import_template.xlsx : real headers + 2 example rows per sheet ----------
tpl = new_wb()

b = add_sheet(tpl, "Batteries", battery_headers())
examples = [
    ["Elecsol-110-001","Elecsol","Carbon Fibre Monoblock","unknown","Unknown",
     110,12.6,4.2,105,92,"Ballinakill","","","24/08/2021","15/03/2021","Bagenalstown","Active"],
    ["Elecsol-110-015","Elecsol","Carbon Fibre Monoblock","unknown","Unknown",
     110,"", "", "", "", "Recycled","Recycled","Tested faulty. Recycled 27-07-2021",
     "N/A","27/07/2021","Maidenhead","Decommissioned"],
]
for r, row in enumerate(examples, 2):
    for i, v in enumerate(row, 1):
        c = b.cell(r, i, v); c.font = EX

m = add_sheet(tpl, "Movements", ["battery_id","hop","from_site","to_site","date_removed","date_installed","reason"])
m.cell(2,1,"Elecsol-110-001"); m.cell(2,2,1); m.cell(2,3,"Bagenalstown")
m.cell(2,4,"Ballinakill"); m.cell(2,5,"15/03/2021"); m.cell(2,6,"24/08/2021"); m.cell(2,7,"Redeployment")
for i in range(1,8): m.cell(2,i).font = EX

add_sheet(tpl, "Readings", [f["key"] for f in config.READING_FIELDS])
o = add_sheet(tpl, "Outages", [f["key"] for f in config.OUTAGE_FIELDS])
o.cell(2,1,"Ballinakill"); o.cell(2,2,"12/05/2024"); o.cell(2,3,"14:30")
o.cell(2,4,3.5); o.cell(2,5,"No"); o.cell(2,6,"Yes"); o.cell(2,7,11.9); o.cell(2,8,"grid trip")
for i in range(1,9): o.cell(2,i).font = EX

s = add_sheet(tpl, "Sites", [f["key"] for f in config.SITE_FIELDS])
for r, sid in enumerate(REAL_SITES, 2):
    s.cell(r,1,sid); s.cell(r,2,""); s.cell(r,3,""); s.cell(r,4,"")

tpl.save("import_template.xlsx")

# ---------- sample_data.xlsx : ~40 batteries to click around ----------
samp = new_wb()
bs = add_sheet(samp, "Batteries", battery_headers())
r = 2
for i in range(1, 41):
    site = random.choice(SITES)
    soh = random.randint(35, 99)
    status = "Decommissioned" if site == "Recycled" else "Active"
    action = {"Recycled":"Recycled","Stock":"Stock"}.get(site, blankish(random.choice(["Keep","Removed"]),0.6))
    row = [f"Elecsol-110-{i:03d}", "Elecsol", "Carbon Fibre Monoblock",
           "unknown", blankish("Unknown",0.3),
           110, blankish(round(random.uniform(11.8,12.9),2),0.3),
           blankish(round(random.uniform(3.5,9),1),0.5),
           blankish(random.randint(90,112),0.4),
           blankish(soh,0.2),
           site, action or "",
           blankish(random.choice(["Leaking electrolyte","Tested faulty","","ok"]),0.5),
           blankish("24/08/2021",0.2), blankish("15/03/2021",0.3),
           blankish(random.choice(REAL_SITES),0.4), status]
    for ci, v in enumerate(row, 1): bs.cell(r, ci, v)
    r += 1

for title, fields in [("Movements",[{"key":k} for k in ["battery_id","hop","from_site","to_site","date_removed","date_installed","reason"]]),
                      ("Readings",config.READING_FIELDS),
                      ("Outages",config.OUTAGE_FIELDS)]:
    add_sheet(samp, title, [f["key"] for f in fields])
ss = add_sheet(samp, "Sites", [f["key"] for f in config.SITE_FIELDS])
for r, sid in enumerate(REAL_SITES, 2):
    ss.cell(r,1,sid)

samp.save("sample_data.xlsx")
print("Wrote import_template.xlsx and sample_data.xlsx")
print("Battery columns:", ", ".join(battery_headers()))
