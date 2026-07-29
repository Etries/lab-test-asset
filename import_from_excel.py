"""
import_from_excel.py  -  One-time (or repeatable) load of a cleaned workbook
into the app's CSV files.

It reads an .xlsx whose sheet names are: Batteries, Movements, Readings,
Outages, Sites  (extra sheets are ignored; missing ones are skipped), and
writes one CSV per sheet into your DATA_DIR.

Usage:
    pip install openpyxl
    python import_from_excel.py path/to/cleaned_workbook.xlsx

WARNING: this OVERWRITES the matching CSVs. A backup of any existing CSV is
taken first (same backups folder the app uses), so you can roll back.
"""
import sys
import os
import openpyxl
import config
import data_store as db

SHEET_TO_TABLE = {
    "Batteries": ("batteries", config.BATTERY_FIELDS),
    "Movements": ("movements", [{"key": k} for k in
                  ["battery_id", "hop", "from_site", "to_site",
                   "date_removed", "date_installed", "reason"]]),
    "Readings":  ("readings", config.READING_FIELDS),
    "Outages":   ("outages", config.OUTAGE_FIELDS),
    "Sites":     ("sites", config.SITE_FIELDS),
}


def import_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet, (table, fields) in SHEET_TO_TABLE.items():
        if sheet not in wb.sheetnames:
            print(f"  - {sheet}: not in workbook, skipped")
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  - {sheet}: empty, skipped")
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        records = []
        for raw in rows[1:]:
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            records.append({headers[i]: ("" if v is None else v)
                            for i, v in enumerate(raw) if i < len(headers)})
        db.save(table, records, headers)
        print(f"  - {sheet}: imported {len(records)} rows -> {table}.csv")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_from_excel.py <cleaned_workbook.xlsx>")
        sys.exit(1)
    src = sys.argv[1]
    if not os.path.exists(src):
        print(f"File not found: {src}")
        sys.exit(1)
    print(f"Importing from {src} into '{config.DATA_DIR}/' ...")
    import_workbook(src)
    print("Done. Start the app with:  python app.py")
