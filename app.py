"""
app.py  -  The web application (routes + logic).

Sections, in order:
   A. Setup + small helpers
   B. Health band + filtering logic  (the battery list)
   C. Pages: battery list, battery detail
   D. Actions: new battery, move battery, add reading
   E. Outages: list + add
   F. Sites: list + detail
   G. Startup

You can read this top to bottom. The things you'd tune (bands, fields, statuses)
are NOT here - they're in config.py.
"""
import datetime
from flask import (Flask, render_template, request, redirect, url_for, flash, abort)

import config
import data_store as db

# ---------------------------------------------------------------------------
# A. SETUP + HELPERS
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = "change-me-if-you-expose-this-app"  # only used for flash messages


def choices_for(spec):
    """Resolve a field's `choices` (which may name a dynamic list) to real values."""
    if isinstance(spec, list):
        return spec
    if spec == "STATUSES":
        return config.STATUSES
    if spec == "BRANDS":
        return sorted({r.get("brand", "") for r in db.load("batteries") if r.get("brand")})
    if spec == "SITES":
        return site_ids()
    if spec == "BATTERIES":
        return sorted(r[config.BATTERY_ID_FIELD] for r in db.load("batteries"))
    return []


def site_ids():
    """All known site ids: from the sites table plus any seen on batteries."""
    ids = {r["site_id"] for r in db.load("sites") if r.get("site_id")}
    ids |= {r.get("current_site", "") for r in db.load("batteries") if r.get("current_site")}
    ids.discard("")
    return sorted(ids)


def to_float(v):
    try:
        return float(str(v).strip())
    except (ValueError, AttributeError):
        return None


@app.context_processor
def inject_globals():
    """Make config values available inside every template."""
    return {"cfg": config, "now": datetime.date.today().isoformat()}


# ---------------------------------------------------------------------------
# B. HEALTH BANDS + FILTERING
# ---------------------------------------------------------------------------
def band_for(row):
    """Return the health band dict for a battery row, based on its SoH%."""
    soh = to_float(row.get(config.SOH_FIELD))
    if soh is None:
        return config.HEALTH_UNKNOWN
    for band in config.HEALTH_BANDS:      # checked top-to-bottom
        if soh >= band["min"]:
            return band
    return config.HEALTH_BANDS[-1]


def filtered_batteries(args):
    """Apply search + facet filters + sort from the query string. Returns rows."""
    rows = db.load("batteries")
    for r in rows:
        r["_band"] = band_for(r)

    # -- search (matches battery id or model, case-insensitive) --
    q = args.get("q", "").strip().lower()
    if q:
        rows = [r for r in rows
                if q in r.get(config.BATTERY_ID_FIELD, "").lower()
                or q in r.get("model", "").lower()]

    # -- facet filters (each may appear multiple times = OR within, AND across) --
    def multi(name):
        return [v for v in args.getlist(name) if v]

    sites = multi("site")
    if sites:
        rows = [r for r in rows if r.get("current_site") in sites]

    statuses = multi("status")
    if statuses:
        rows = [r for r in rows if r.get("status") in statuses]

    brands = multi("brand")
    if brands:
        rows = [r for r in rows if r.get("brand") in brands]

    bands = multi("health")
    if bands:
        rows = [r for r in rows if r["_band"]["key"] in bands]

    # -- sort --
    sort = args.get("sort", "battery_id")
    reverse = args.get("dir", "asc") == "desc"
    if sort == "soh":
        rows.sort(key=lambda r: (to_float(r.get(config.SOH_FIELD)) is None,
                                 to_float(r.get(config.SOH_FIELD)) or 0), reverse=reverse)
    else:
        rows.sort(key=lambda r: r.get(sort, ""), reverse=reverse)
    return rows


# ---------------------------------------------------------------------------
# C. PAGES: battery list + detail
# ---------------------------------------------------------------------------
@app.route("/")
def batteries():
    rows = filtered_batteries(request.args)
    return render_template(
        "batteries.html",
        rows=rows,
        fields=[f for f in config.BATTERY_FIELDS if f.get("in_table")],
        all_sites=site_ids(),
        all_statuses=config.STATUSES,
        all_brands=choices_for("BRANDS"),
        args=request.args,
        total=len(db.load("batteries")),
    )


@app.route("/battery/<bid>")
def battery_detail(bid):
    rows = db.load("batteries")
    match = next((r for r in rows if r.get(config.BATTERY_ID_FIELD) == bid), None)
    if not match:
        abort(404)
    match["_band"] = band_for(match)
    moves = [m for m in db.load("movements") if m.get("battery_id") == bid]
    moves.sort(key=lambda m: m.get("hop", ""))
    readings = [x for x in db.load("readings") if x.get("battery_id") == bid]
    readings.sort(key=lambda x: x.get("reading_date", ""))
    return render_template("battery_detail.html", b=match, fields=config.BATTERY_FIELDS,
                           moves=moves, readings=readings)


# ---------------------------------------------------------------------------
# D. ACTIONS: new battery, move battery, add reading
# ---------------------------------------------------------------------------
@app.route("/battery/new", methods=["GET", "POST"])
def battery_new():
    if request.method == "POST":
        row = {}
        for f in config.BATTERY_FIELDS:
            row[f["key"]] = request.form.get(f["key"], "").strip()
        # validation: required fields + unique id
        for f in config.BATTERY_FIELDS:
            if f.get("required") and not row[f["key"]]:
                flash(f"{f['label']} is required.", "error")
                return render_template("battery_form.html", fields=config.BATTERY_FIELDS,
                                       row=row, choices_for=choices_for, mode="new")
        existing = {r[config.BATTERY_ID_FIELD] for r in db.load("batteries")}
        if row[config.BATTERY_ID_FIELD] in existing:
            flash(f"A battery with ID '{row[config.BATTERY_ID_FIELD]}' already exists.", "error")
            return render_template("battery_form.html", fields=config.BATTERY_FIELDS,
                                   row=row, choices_for=choices_for, mode="new")
        db.append("batteries", row, config.BATTERY_FIELDS)
        flash(f"Added battery {row[config.BATTERY_ID_FIELD]}.", "ok")
        return redirect(url_for("battery_detail", bid=row[config.BATTERY_ID_FIELD]))

    blank = {f["key"]: "" for f in config.BATTERY_FIELDS}
    blank["status"] = "Active"
    return render_template("battery_form.html", fields=config.BATTERY_FIELDS,
                           row=blank, choices_for=choices_for, mode="new")


@app.route("/battery/<bid>/move", methods=["GET", "POST"])
def battery_move(bid):
    rows = db.load("batteries")
    b = next((r for r in rows if r.get(config.BATTERY_ID_FIELD) == bid), None)
    if not b:
        abort(404)

    if request.method == "POST":
        to_site = request.form.get("to_site", "").strip()
        reason = request.form.get("reason", "").strip()
        date_removed = request.form.get("date_removed", "").strip()
        date_installed = request.form.get("date_installed", "").strip()
        remarks = request.form.get("remarks", "").strip()
        if not to_site:
            flash("Destination site is required.", "error")
            return redirect(url_for("battery_move", bid=bid))

        from_site = b.get("current_site", "")
        # 1. record the movement
        existing_hops = [int(m.get("hop", 0) or 0) for m in db.load("movements")
                         if m.get("battery_id") == bid]
        next_hop = (max(existing_hops) + 1) if existing_hops else 1
        db.append("movements", {
            "battery_id": bid, "hop": next_hop, "from_site": from_site,
            "to_site": to_site, "date_removed": date_removed,
            "date_installed": date_installed, "reason": reason,
        }, [{"key": k} for k in ["battery_id", "hop", "from_site", "to_site",
                                 "date_removed", "date_installed", "reason"]])

        # 2. update the battery's current state (this is what keeps them in sync)
        b["current_site"] = to_site
        b["last_moved_date"] = date_removed or datetime.date.today().isoformat()
        b["date_fitted"] = date_installed
        if remarks:
            b["remarks"] = remarks
        forced = next((r["sets_status"] for r in config.MOVEMENT_REASONS
                       if r["reason"] == reason), None)
        if forced:
            b["status"] = forced
        headers = db.load_headers("batteries", config.BATTERY_FIELDS)
        db.save("batteries", rows, headers)

        flash(f"Moved {bid}: {from_site or '—'} → {to_site}.", "ok")
        return redirect(url_for("battery_detail", bid=bid))

    return render_template("battery_move.html", b=b, sites=site_ids(),
                           reasons=config.MOVEMENT_REASONS)


@app.route("/battery/<bid>/reading", methods=["POST"])
def add_reading(bid):
    rows = db.load("batteries")
    b = next((r for r in rows if r.get(config.BATTERY_ID_FIELD) == bid), None)
    if not b:
        abort(404)
    reading = {
        "battery_id": bid,
        "reading_date": request.form.get("reading_date", "").strip(),
        "voltage": request.form.get("voltage", "").strip(),
        "soh_pct": request.form.get("soh_pct", "").strip(),
        "source": request.form.get("source", "Manual").strip(),
    }
    db.append("readings", reading, config.READING_FIELDS)
    # update the battery's latest values too
    if reading["voltage"]:
        b["tested_voltage"] = reading["voltage"]
    if reading["soh_pct"]:
        b["soh_pct"] = reading["soh_pct"]
    db.save("batteries", rows, db.load_headers("batteries", config.BATTERY_FIELDS))
    flash("Reading recorded.", "ok")
    return redirect(url_for("battery_detail", bid=bid))


# ---------------------------------------------------------------------------
# E. OUTAGES
# ---------------------------------------------------------------------------
@app.route("/outages")
def outages():
    rows = db.load("outages")
    rows.sort(key=lambda r: r.get("outage_date", ""), reverse=True)
    site = request.args.get("site", "")
    if site:
        rows = [r for r in rows if r.get("site_id") == site]
    return render_template("outages.html", rows=rows, fields=config.OUTAGE_FIELDS,
                           all_sites=site_ids(), args=request.args)


@app.route("/outages/new", methods=["GET", "POST"])
def outage_new():
    if request.method == "POST":
        row = {f["key"]: request.form.get(f["key"], "").strip() for f in config.OUTAGE_FIELDS}
        for f in config.OUTAGE_FIELDS:
            if f.get("required") and not row[f["key"]]:
                flash(f"{f['label']} is required.", "error")
                return render_template("outage_form.html", fields=config.OUTAGE_FIELDS,
                                       row=row, choices_for=choices_for)
        db.append("outages", row, config.OUTAGE_FIELDS)
        flash("Outage recorded.", "ok")
        return redirect(url_for("outages"))
    blank = {f["key"]: "" for f in config.OUTAGE_FIELDS}
    return render_template("outage_form.html", fields=config.OUTAGE_FIELDS,
                           row=blank, choices_for=choices_for)


# ---------------------------------------------------------------------------
# F. SITES
# ---------------------------------------------------------------------------
@app.route("/sites")
def sites():
    bats = db.load("batteries")
    for r in bats:
        r["_band"] = band_for(r)
    summary = []
    for sid in site_ids():
        here = [r for r in bats if r.get("current_site") == sid]
        by_brand = {}
        for r in here:
            by_brand[r.get("brand", "—")] = by_brand.get(r.get("brand", "—"), 0) + 1
        summary.append({
            "site_id": sid,
            "count": len(here),
            "by_brand": by_brand,
            "needs_replace": sum(1 for r in here if r["_band"]["key"] == "replace"),
        })
    return render_template("sites.html", summary=summary)


@app.route("/site/<sid>")
def site_detail(sid):
    bats = [r for r in db.load("batteries") if r.get("current_site") == sid]
    for r in bats:
        r["_band"] = band_for(r)
    bats.sort(key=lambda r: r.get("brand", ""))
    site_row = next((s for s in db.load("sites") if s.get("site_id") == sid), {"site_id": sid})
    site_outages = [o for o in db.load("outages") if o.get("site_id") == sid]
    site_outages.sort(key=lambda o: o.get("outage_date", ""), reverse=True)
    return render_template("site_detail.html", sid=sid, site=site_row,
                           bats=bats, outages=site_outages,
                           fields=[f for f in config.BATTERY_FIELDS if f.get("in_table")])


# ---------------------------------------------------------------------------
# G. IMPORT / EXPORT (the workbook)
# ---------------------------------------------------------------------------
@app.route("/data")
def data_page():
    """Import/export hub: upload a workbook, or download the current one."""
    import os
    wb_exists = os.path.exists(config.WORKBOOK_FILE)
    counts = {t: len(db.load(t)) for t in
              ["batteries", "movements", "readings", "outages", "sites"]}
    return render_template("data.html", wb_exists=wb_exists, counts=counts)


@app.route("/data/export")
def data_export():
    """(Re)build the workbook and send it as a download."""
    from flask import send_file
    path = db.export_workbook()
    return send_file(path, as_attachment=True,
                     download_name="battery_inventory.xlsx")


@app.route("/data/import", methods=["POST"])
def data_import():
    """Read an uploaded .xlsx back into the CSVs (hand-edits come back in here).
    A backup of current data is taken automatically before overwriting."""
    import os
    file = request.files.get("workbook")
    if not file or not file.filename:
        flash("Please choose a file to import.", "error")
        return redirect(url_for("data_page"))
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload an Excel .xlsx file.", "error")
        return redirect(url_for("data_page"))

    tmp_path = os.path.join(config.DATA_DIR, "_uploaded.xlsx")
    os.makedirs(config.DATA_DIR, exist_ok=True)
    file.save(tmp_path)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        sheet_map = {
            "Batteries": ("batteries", config.BATTERY_FIELDS),
            "Movements": ("movements", [{"key": k} for k in
                          ["battery_id", "hop", "from_site", "to_site",
                           "date_removed", "date_installed", "reason"]]),
            "Readings": ("readings", config.READING_FIELDS),
            "Outages": ("outages", config.OUTAGE_FIELDS),
            "Sites": ("sites", config.SITE_FIELDS),
        }
        summary = []
        for sheet, (table, fields) in sheet_map.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            records = []
            for raw in rows[1:]:
                if all(c is None or str(c).strip() == "" for c in raw):
                    continue
                records.append({headers[i]: ("" if v is None else v)
                                for i, v in enumerate(raw) if i < len(headers)})
            db.save(table, records, headers)
            summary.append(f"{sheet}: {len(records)}")
        flash("Imported — " + ", ".join(summary) + ". The workbook is now in sync.", "ok")
    except Exception as e:
        flash(f"Import failed: {e}", "error")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return redirect(url_for("data_page"))


# ---------------------------------------------------------------------------
# H. STARTUP
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import webbrowser
    import threading
    url = f"http://127.0.0.1:{config.PORT}"
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"\n  {config.APP_TITLE} running at {url}\n  Press Ctrl+C to stop.\n")
    app.run(port=config.PORT, debug=True, use_reloader=False)
