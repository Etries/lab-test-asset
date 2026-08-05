"""
data_store.py  -  Reads and writes the CSV files that hold all app data.

You normally won't edit this. It gives the rest of the app a simple way to load
a table (list of dict rows) and save it back safely:

  - saves are ATOMIC (write to a temp file, then rename) so a crash mid-save
    can't leave you with a half-written, corrupt CSV.
  - every save first drops a timestamped backup in the backups folder.

Each "table" is one CSV file: batteries.csv, movements.csv, readings.csv,
outages.csv, sites.csv
"""
import csv
import os
import shutil
import datetime
import config


def _path(table):
    return os.path.join(config.DATA_DIR, f"{table}.csv")


def _ensure_dirs():
    os.makedirs(config.DATA_DIR, exist_ok=True)
    os.makedirs(config.BACKUP_DIR, exist_ok=True)


def load(table):
    """Return a list of row dicts for the given table (empty list if no file)."""
    path = _path(table)
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def load_headers(table, default_fields):
    """Return the CSV's column names, or the configured defaults if no file yet."""
    path = _path(table)
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if headers:
                return headers
    return [f["key"] for f in default_fields]


def _backup(table):
    path = _path(table)
    if not os.path.exists(path):
        return
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = os.path.join(config.BACKUP_DIR, f"{table}.{stamp}.csv")
    shutil.copy2(path, dest)
    _prune_backups(table)


def _prune_backups(table):
    files = sorted(
        (f for f in os.listdir(config.BACKUP_DIR) if f.startswith(f"{table}.")),
        reverse=True,
    )
    for old in files[config.KEEP_BACKUPS:]:
        try:
            os.remove(os.path.join(config.BACKUP_DIR, old))
        except OSError:
            pass


def _refresh_workbook():
    """Rewrite the shared .xlsx snapshot. Guarded so a workbook problem (e.g. the
    file is open in Excel and locked) never blocks the CSV save that just
    succeeded — it just prints a note and moves on."""
    if not getattr(config, "SYNC_WORKBOOK", True):
        return
    try:
        export_workbook()
    except PermissionError:
        print("  [workbook] Could not update the .xlsx (is it open in Excel? "
              "the CSV data is saved regardless).")
    except Exception as e:
        print(f"  [workbook] Skipped .xlsx refresh: {e}")


def save(table, rows, headers):
    """Write rows (list of dicts) to the table's CSV, atomically, with a backup.
    Also refreshes the shared workbook snapshot so the .xlsx stays current."""
    _ensure_dirs()
    _backup(table)
    path = _path(table)
    tmp = path + ".tmp"
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in headers})
    os.replace(tmp, path)  # atomic on the same filesystem
    _refresh_workbook()


def append(table, row, default_fields):
    """Convenience: load, add one row, save. Keeps existing headers."""
    rows = load(table)
    headers = load_headers(table, default_fields)
    # make sure any brand-new keys still get a column
    for k in row:
        if k not in headers:
            headers.append(k)
    rows.append(row)
    save(table, rows, headers)
    return rows


# ---------------------------------------------------------------------------
# WORKBOOK SYNC
# The app keeps an .xlsx copy of everything in sync with the CSVs, so anyone
# can open it to read (or take offline). This is written on every change.
# The workbook is a SNAPSHOT for reading/sharing; the app remains the editor.
# To feed hand-edits back in, use the Import button (which reads a workbook
# back into the CSVs).
# ---------------------------------------------------------------------------
TABLES = ["batteries", "movements", "readings", "outages", "sites"]
SHEET_NAMES = {
    "batteries": "Batteries", "movements": "Movements", "readings": "Readings",
    "outages": "Outages", "sites": "Sites",
}


def export_workbook(path=None):
    """Write all tables into one .xlsx (one sheet per table)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    path = path or getattr(config, "WORKBOOK_FILE", os.path.join(config.DATA_DIR, "battery_inventory.xlsx"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    default_fields = {
        "batteries": config.BATTERY_FIELDS,
        "outages": config.OUTAGE_FIELDS,
        "readings": config.READING_FIELDS,
        "sites": config.SITE_FIELDS,
        "movements": [{"key": k} for k in
                      ["battery_id", "hop", "from_site", "to_site",
                       "date_removed", "date_installed", "reason"]],
    }
    for table in TABLES:
        ws = wb.create_sheet(SHEET_NAMES[table])
        headers = load_headers(table, default_fields[table])
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h); c.font = hdr_font; c.fill = hdr_fill
        for r, row in enumerate(load(table), 2):
            for i, h in enumerate(headers, 1):
                ws.cell(r, i, row.get(h, ""))
        ws.freeze_panes = "A2"
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)   # atomic
    return path
